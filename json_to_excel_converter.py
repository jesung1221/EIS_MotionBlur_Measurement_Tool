import pandas as pd
import json
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import os

def convert_json_to_excel(json_file_path, output_file):
    # Load the JSON file
    with open(json_file_path, 'r') as file:
        data = json.load(file)

    # Convert JSON data to a pandas DataFrame
    df = pd.DataFrame.from_dict(data, orient='index')

    # Check and handle missing degree_of_eis_fix
    if 'degree_of_eis_fix' not in df.columns:
        df['degree_of_eis_fix'] = np.nan
        print("Warning: 'degree_of_eis_fix' not found in data. Setting to NaN.")
    else:
        # Calculate the Suppression Ratio only if degree_of_eis_fix exists
        df['Suppression Ratio'] = 20 * np.log10(20 / df['degree_of_eis_fix'].abs().replace(0, np.nan))

    # Save the DataFrame to an Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=True, sheet_name='Video Info Summary')

    # Load the workbook and select the sheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Adjust the column widths
    for column in ws.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_name].width = adjusted_width

    # Save the workbook
    wb.save(output_file)

    print(f'Successfully saved the summary to {output_file}')
    
    # Create plots and insert them into the Excel sheet
    create_plots(df, output_file)


def create_plots(df, output_file):
    # Extract unique camera devices
    camera_devices = df['camera_device'].unique()

    # First plot: degree_of_eis_fix vs rpm
    fig1, ax1 = plt.subplots(figsize=(10, 6))
    for device in camera_devices:
        device_df = df[df['camera_device'] == device]
        ax1.plot(device_df['rpm'], device_df['degree_of_eis_fix'], marker='o', label=device)
    ax1.set_xlabel('RPM')
    ax1.set_ylabel('Degree of EIS Fix')
    ax1.set_title('Degree of EIS Fix vs RPM')
    ax1.legend()
    ax1.grid(True)
    plot1_path = output_file.replace('.xlsx', '_degree_of_eis_fix_plot.png')
    fig1.savefig(plot1_path)
    plt.close(fig1)

    # Second plot: Suppression Ratio vs rpm
    fig2, ax2 = plt.subplots(figsize=(10, 6))
    for device in camera_devices:
        device_df = df[df['camera_device'] == device]
        ax2.plot(device_df['rpm'], device_df['Suppression Ratio'], marker='o', label=device)
    ax2.set_xlabel('RPM')
    ax2.set_ylabel('Suppression Ratio (dB)')
    ax2.set_title('Suppression Ratio vs RPM')
    ax2.legend()
    ax2.grid(True)
    plot2_path = output_file.replace('.xlsx', '_suppression_ratio_plot.png')
    fig2.savefig(plot2_path)
    plt.close(fig2)

    # Third plot: motion blur vs rpm
    fig3, ax3 = plt.subplots(figsize=(10, 6))
    for device in camera_devices:
        device_df = df[df['camera_device'] == device]
        ax3.plot(device_df['rpm'], device_df['motion_blur'], marker='o', label=device)
    ax3.set_xlabel('RPM')
    ax3.set_ylabel('Motion Blur')
    ax3.set_title('Motion Blur vs RPM')
    ax3.legend()
    ax3.grid(True)
    plot3_path = output_file.replace('.xlsx', '_motion_blur_plot.png')
    fig3.savefig(plot3_path)
    plt.close(fig3)

    # Load the workbook and select the sheet
    wb = load_workbook(output_file)
    ws = wb.active

    # Insert plots into the Excel sheet
    img1 = Image(plot1_path)
    img2 = Image(plot2_path)
    img3 = Image(plot3_path)
    ws.add_image(img1, 'J1')
    ws.add_image(img2, 'J20')
    ws.add_image(img3, 'J39')

    # Save the workbook with the images
    wb.save(output_file)

    # Remove the temporary plot images
    os.remove(plot1_path)
    os.remove(plot2_path)
    os.remove(plot3_path)

    print('Plots have been inserted into the Excel sheet and saved.')